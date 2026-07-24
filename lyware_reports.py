"""
LYWARE — report export layer.

Turns the read-only analytics in `lyware.py` into print-ready LibreOffice Calc
spreadsheets (.xlsx, which Calc opens and recalculates natively). One generic
sheet writer drives every preset, so adding a report is just a column map + a
row source. `openpyxl` is imported lazily so the rest of the app never depends
on it; if it is missing the GUI shows a friendly "pip install openpyxl" message.

Public API:
    REPORTS                      -> ordered list of (key, label, description)
    report_columns(key)          -> the column layout a given report will produce
    build_report(conn, key, filt, out_path)            -> writes a preset report
    build_custom_report(conn, source, columns, filt, out_path, title)  -> custom
"""
from datetime import datetime

import lyware as L

# Lazy openpyxl handle; resolved on first use.
_OPX = None


def _openpyxl():
    global _OPX
    if _OPX is None:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        _OPX = dict(wb=openpyxl.Workbook, Font=Font, Fill=PatternFill, Align=Alignment,
                    Border=Border, Side=Side, col=get_column_letter)
    return _OPX


def openpyxl_available():
    try:
        _openpyxl(); return True
    except Exception:
        return False


# --- report catalogue -------------------------------------------------------
# Each entry: key -> (label, description, source-fn-name, [(header, dict_key, kind)])
# kind: "text" | "int" | "money" (money columns get a currency format + a SUM total)
_DEF = {
    "transactions": (
        "Transactions", "Every ledger movement, account and category — the accounting workbook.",
        "transactions_report",
        [("ID", "trnsid", "int"), ("Date", "date", "text"), ("Time", "time", "text"),
         ("Account", "account_name", "text"), ("Type", "type", "text"),
         ("Category", "category", "text"), ("Amount", "amount", "money"),
         ("Currency", "currency", "text")]),
    "sales": (
        "Sales", "Per-unit sale price, cost and realized profit.",
        "sales_report",
        [("Order", "order", "int"), ("Buyer", "buyer", "text"), ("Item", "item", "text"),
         ("Sale price", "sale_price", "money"), ("Cost", "cost", "money"),
         ("Additional", "additional", "money"), ("Profit", "profit", "money"),
         ("Status", "status", "text"), ("Committed", "committed", "text"),
         ("Finalized", "finalized", "text")]),
    "inventory": (
        "Inventory", "Every unit the shop owns or has owned, with stage and cost.",
        "inventory_report",
        [("Unit", "lywrid", "int"), ("Item", "item", "text"), ("Status", "status", "text"),
         ("Stage", "stage", "text"), ("Cost", "cost", "money"), ("Purchased", "purchased", "text"),
         ("Entered stock", "entered", "text"), ("Method", "method", "text")]),
    "logistics": (
        "Logistics", "The full inbound timeline of each unit, milestone by milestone.",
        "logistics_report",
        [("Unit", "lywrid", "int"), ("Item", "item", "text"), ("Status", "status", "text"),
         ("Method", "method", "text"), ("Purchased", "purchased", "text"),
         ("US warehouse", "us_warehouse", "text"), ("LY warehouse", "libya_warehouse", "text"),
         ("Local sent", "local_sent", "text"), ("Local office", "local_office", "text"),
         ("Picked up", "picked_up", "text")]),
    "catalogue": (
        "Catalogue performance", "Listing / purchase / sale aggregates per catalogue item.",
        "catalog_performance",
        [("Item", "item", "text"), ("Category", "category", "text"),
         ("Local value (LYD, 90d active)", "local_value_avg", "money"),
         ("Local listings (n)", "local_value_n", "int"),
         ("Times listed", "times_listed", "int"), ("Avg list price (LYD @ market rate)", "avg_listing_price", "money"),
         ("Qty purchased", "qty_purchased", "int"), ("Avg buy cost", "avg_purchase_cost", "money"),
         ("Qty sold", "qty_sold", "int"), ("Avg sale price", "avg_sale_price", "money"),
         ("Sale volume", "sale_volume", "money"), ("Avg margin", "avg_margin", "money")]),
    "fx": (
        "FX / USD market", "Every USD acquisition with the LYD-per-USD rate, over time.",
        "fx_report",
        [("Date", "date", "text"), ("Account", "account", "text"), ("USD", "usd", "money"),
         ("LYD cost", "lyd_cost", "money"), ("Rate (LYD/USD)", "rate", "money"),
         ("Source", "source", "text")]),
    "accounts": (
        "Accounts", "Balance snapshot of every account.",
        "accounts_report",
        [("ID", "acctid", "int"), ("Account", "account", "text"), ("Type", "type", "text"),
         ("Balance LYD", "lyd", "money"), ("Balance USD", "usd", "money"),
         ("Created", "created", "text")]),
    "losses": (
        "Losses & Returns", "Cancelled, written-off and returned units with cost, recovery and net.",
        "losses_report",
        [("Unit", "lywrid", "int"), ("Item", "item", "text"), ("Outcome", "status", "text"),
         ("Date", "date", "text"), ("Sunk cost", "cost", "money"), ("Recovery", "recovery", "money"),
         ("Net", "net", "money"), ("Note", "note", "text")]),
}

REPORTS = [(k, v[0], v[1]) for k, v in _DEF.items()]


def report_columns(key):
    return [(h, kind) for h, _, kind in _DEF[key][3]]


# --- workbook writer --------------------------------------------------------
def _filter_caption(filt):
    f = filt or {}
    bits = []
    if f.get("date_from") or f.get("date_to"):
        bits.append(f"Dates {f.get('date_from', '…')} → {f.get('date_to', '…')}")
    for k, label in (("platform", "Platform"), ("category", "Category"),
                     ("status", "Status"), ("acctid", "Account #"), ("catid", "Catalogue #")):
        if f.get(k):
            bits.append(f"{label}: {f[k]}")
    return " | ".join(bits) if bits else "No filters (all data)"


def _write_sheet(wb, sheet_title, columns, dict_rows, filt, report_label, first=True):
    O = _openpyxl()
    Font, Fill, Align, Border, Side, col = (O["Font"], O["Fill"], O["Align"],
                                            O["Border"], O["Side"], O["col"])
    ws = wb.active if first else wb.create_sheet()
    safe = sheet_title
    for ch in '/\\?*[]:':
        safe = safe.replace(ch, "-")
    ws.title = safe[:31]
    blue = "1F4E79"
    thin = Side(style="thin", color="D0D0D0")
    border = Border(bottom=thin, right=thin)

    ws["A1"] = "LYWARE"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=blue)
    ws["A2"] = f"{report_label} report"
    ws["A2"].font = Font(name="Arial", bold=True, size=12)
    ws["A3"] = _filter_caption(filt)
    ws["A3"].font = Font(name="Arial", italic=True, size=9, color="666666")
    ws["A4"] = "Generated " + datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"].font = Font(name="Arial", italic=True, size=9, color="666666")

    head_row = 6
    headers = [h for h, _ in columns]
    money_idx = [i for i, (_, kind) in enumerate(columns) if kind == "money"]
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=head_row, column=j, value=h)
        cell.font = Font(name="Arial", bold=True, color="FFFFFF")
        cell.fill = Fill("solid", fgColor=blue)
        cell.alignment = Align(horizontal="center", vertical="center")
        cell.border = border

    r = head_row + 1
    for row in dict_rows:
        for j, (h, kind) in enumerate(columns, start=1):
            v = row.get(_key_for(report_label, h), row.get(h))
            val = _coerce(v, kind)
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = Font(name="Arial")
            cell.border = border
            if kind == "money":
                cell.number_format = '#,##0.00;(#,##0.00);"-"'
            elif kind == "int":
                cell.alignment = Align(horizontal="center")
        r += 1
    last = r - 1

    # totals row (Calc recalculates the SUM formulas on open)
    if last >= head_row + 1 and money_idx:
        ws.cell(row=r, column=1, value="TOTAL").font = Font(name="Arial", bold=True)
        for i in money_idx:
            c = col(i + 1)
            cell = ws.cell(row=r, column=i + 1, value=f"=SUM({c}{head_row + 1}:{c}{last})")
            cell.font = Font(name="Arial", bold=True)
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
            cell.border = Border(top=Side(style="thin", color="808080"))

    for j, h in enumerate(headers, start=1):
        width = max(11, min(34, len(h) + 4,
                            max([len(str(row.get(_key_for(report_label, h), ""))) for row in dict_rows] + [len(h)]) + 3))
        ws.column_dimensions[col(j)].width = width
    ws.freeze_panes = ws.cell(row=head_row + 1, column=1)
    return ws


# header label -> dict key resolution is done by position in the preset; for the
# generic path we pass dict rows already keyed by the canonical key, so this maps
# a header back to the canonical key for the active report.
_LABEL_KEYMAP = {}


def _key_for(report_label, header):
    return _LABEL_KEYMAP.get((report_label, header), header)


def _coerce(v, kind):
    if v is None or v == "":
        return None
    if kind in ("money", "int"):
        try:
            return float(v)
        except (TypeError, ValueError):
            return str(v)
    return str(v)


def build_report(conn, key, filt, out_path):
    """Write a preset report to `out_path` (.xlsx). Returns the path."""
    label, _desc, source, columns = _DEF[key]
    for h, dk, _ in columns:               # register header->key for this preset
        _LABEL_KEYMAP[(label, h)] = dk
    rows = getattr(L, source)(conn, filt)
    rows = [{k: (str(v) if hasattr(v, "is_nan") else v) for k, v in r.items()} for r in rows]
    O = _openpyxl()
    wb = O["wb"]()
    _write_sheet(wb, label, [(h, kind) for h, _, kind in columns], rows, filt, label)
    wb.save(out_path)
    return out_path


def build_summary_report(conn, filt, out_path):
    """A one-page Financial Summary (revenue / expenses / net) plus the transaction
    detail behind it — the closest thing to a printable P&L statement."""
    fs = L.financial_summary(conn, filt)
    O = _openpyxl()
    Font, Fill, Align = O["Font"], O["Fill"], O["Align"]
    wb = O["wb"]()
    ws = wb.active
    ws.title = "Summary"
    blue = "1F4E79"
    ws["A1"] = "LYWARE"; ws["A1"].font = Font(name="Arial", bold=True, size=16, color=blue)
    ws["A2"] = "Financial summary"; ws["A2"].font = Font(name="Arial", bold=True, size=12)
    ws["A3"] = _filter_caption(filt); ws["A3"].font = Font(name="Arial", italic=True, size=9, color="666666")
    ws["A4"] = "Generated " + datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"].font = Font(name="Arial", italic=True, size=9, color="666666")
    rows = [("Revenue", float(fs["revenue"])), ("Expenses", -float(fs["expense"]))]
    for tp, amt in sorted(fs["expense_by_type"].items()):
        rows.append((f"    {tp}", -float(amt)))
    rows.append(("FX gain / (loss)", float(fs["fx_gain_loss"])))
    rows.append(("NET", float(fs["net"])))
    rows.append(("", None))
    rows.append(("Capital deposited", float(fs["deposits"])))
    rows.append(("Capital withdrawn", -float(fs["withdrawals"])))
    r = 6
    for label, amt in rows:
        ws.cell(row=r, column=1, value=label).font = Font(
            name="Arial", bold=label in ("Revenue", "Expenses", "NET"))
        if amt is not None:
            cell = ws.cell(row=r, column=2, value=amt)
            cell.number_format = '#,##0.00;(#,##0.00);"-"'
            cell.font = Font(name="Arial", bold=label == "NET")
        r += 1
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18
    # detail sheet
    label = "Transactions"
    for h, dk, _ in _DEF["transactions"][3]:
        _LABEL_KEYMAP[(label, h)] = dk
    detail = L.transactions_report(conn, filt)
    _write_sheet(wb, "Detail", [(h, kind) for h, _, kind in _DEF["transactions"][3]],
                 detail, filt, label, first=False)
    wb.save(out_path)
    return out_path


def build_custom_report(conn, key, chosen_headers, filt, out_path):
    """Custom report: a preset's source filtered to a user-chosen subset of columns."""
    label, _desc, source, columns = _DEF[key]
    keep = [(h, kind) for (h, dk, kind) in columns if h in chosen_headers]
    for h, dk, _ in columns:
        _LABEL_KEYMAP[(label, h)] = dk
    rows = getattr(L, source)(conn, filt)
    O = _openpyxl()
    wb = O["wb"]()
    _write_sheet(wb, label, keep, rows, filt, label)
    wb.save(out_path)
    return out_path
